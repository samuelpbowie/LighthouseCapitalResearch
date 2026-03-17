"""
extract_okta_data.py
Reads Okta Model.xlsm and outputs equity-research/data/okta_data.json
Run from the repo root: python scripts/extract_okta_data.py
"""

import json
import os
import openpyxl
from datetime import datetime

MODEL_PATH  = os.path.join(os.path.dirname(__file__), '..', 'model_files', 'Okta Model.xlsm')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'equity-research', 'data', 'okta_data.json')

def r(v, decimals=2):
    """Round a float; return None if falsy."""
    if v is None:
        return None
    try:
        return round(float(v), decimals)
    except (TypeError, ValueError):
        return None

def pct(v, decimals=2):
    """Convert decimal to percentage, rounded."""
    if v is None:
        return None
    try:
        return round(float(v) * 100, decimals)
    except (TypeError, ValueError):
        return None

# ── Load workbook ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(MODEL_PATH, read_only=True, keep_vba=True, data_only=True)

# ── 1. Current stock price from Evaluation ──────────────────────────────────
eval_ws = wb['Evaluation']
eval_rows = list(eval_ws.iter_rows(values_only=True))
current_price = None
for row in eval_rows:
    if row[1] == 'Current Stock Price:':
        current_price = r(row[2])
        break

# ── 2. Consolidated Output metrics ──────────────────────────────────────────
co_ws   = wb['Consolidated Output']
co_rows = list(co_ws.iter_rows(values_only=True))

# Row 2 is the header; cols 2-10 are years (datetime objects)
header_row = co_rows[1]  # 0-indexed row 1 = spreadsheet row 2
LAST_HIST_YEAR = 2026  # last reported fiscal year
years = []
for cell in header_row[2:]:
    if isinstance(cell, datetime):
        fy = cell.year
        label = f'FY{fy}' if fy <= LAST_HIST_YEAR else f'FY{fy}E'
        years.append(label)
    elif cell is not None:
        years.append(str(cell))

# Build metric lookup: {metric_name: {scenario: [values...]}}
metrics_raw = {}
for row in co_rows[2:]:  # data starts at row 3
    if row[0] is None:
        continue
    metric   = row[0]
    scenario = row[1]
    values   = list(row[2:])
    if metric not in metrics_raw:
        metrics_raw[metric] = {}
    metrics_raw[metric][scenario] = values

SCENARIOS = ['Base', 'Pessimistic', 'Optimistic', 'Weighted Average']

def extract_metric(name, transform=r):
    out = {}
    if name not in metrics_raw:
        return out
    for s in SCENARIOS:
        if s in metrics_raw[name]:
            out[s] = [transform(v) for v in metrics_raw[name][s]]
    return out

def compute_ebitda_margin():
    """Compute EBITDA / Revenue directly to avoid any pre-calc issues."""
    out = {}
    for s in SCENARIOS:
        rev_vals   = metrics_raw.get('Revenue', {}).get(s, [])
        ebitda_vals= metrics_raw.get('EBITDA',  {}).get(s, [])
        out[s] = [
            r(e / rv * 100, 1) if (rv and e is not None) else None
            for e, rv in zip(ebitda_vals, rev_vals)
        ]
    return out

def compute_fcf_margin():
    out = {}
    for s in SCENARIOS:
        rev_vals = metrics_raw.get('Revenue', {}).get(s, [])
        fcf_vals = metrics_raw.get('FCF before Buybacks', {}).get(s, [])
        out[s] = [
            r(f / rv * 100, 1) if (rv and f is not None) else None
            for f, rv in zip(fcf_vals, rev_vals)
        ]
    return out

metrics = {
    'Revenue':     extract_metric('Revenue'),
    'RevenueCagr': extract_metric('Revenue CAGR', pct),
    'EBITDA':      extract_metric('EBITDA'),
    'EBITDAMargin':compute_ebitda_margin(),
    'EPS':         extract_metric('EPS'),
    'FCF':         extract_metric('FCF before Buybacks'),
    'FCFMargin':   compute_fcf_margin(),
}

# ── 3. Three-statement model data (gross profit / margins) ───────────────────
def parse_model_sheet(sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    data = {}
    for row in rows:
        if row[0] is not None:
            data[row[0]] = list(row[1:])
    return data

base_model = parse_model_sheet('Model Base')
pess_model = parse_model_sheet('Model Pess.')
opt_model  = parse_model_sheet('Model Opt.')

def get_row(model, key, transform=r):
    vals = model.get(key, [None]*9)
    return [transform(v) for v in vals]

def wa_row(base_vals, pess_vals, opt_vals):
    """Compute Weighted Average = 50% Base + 15% Pess + 35% Opt."""
    result = []
    for b, p, o in zip(base_vals, pess_vals, opt_vals):
        if b is None and p is None and o is None:
            result.append(None)
        else:
            result.append(r((b or 0)*0.5 + (p or 0)*0.15 + (o or 0)*0.35))
    return result

gp_base = get_row(base_model, 'Gross Profit')
gp_pess = get_row(pess_model, 'Gross Profit')
gp_opt  = get_row(opt_model,  'Gross Profit')
rev_base = get_row(base_model, 'Total Revenues:')
rev_pess = get_row(pess_model, 'Total Revenues:')
rev_opt  = get_row(opt_model,  'Total Revenues:')
ni_base = get_row(base_model, 'Net Income:')
ni_pess = get_row(pess_model, 'Net Income:')
ni_opt  = get_row(opt_model,  'Net Income:')

gp_wa  = wa_row(gp_base, gp_pess, gp_opt)
rev_wa = wa_row(rev_base, rev_pess, rev_opt)
ni_wa  = wa_row(ni_base, ni_pess, ni_opt)

om_base = [pct(v) for v in base_model.get('Operating Margin:', [None]*9)]
om_pess = [pct(v) for v in pess_model.get('Operating Margin:', [None]*9)]
om_opt  = [pct(v) for v in opt_model.get('Operating Margin:',  [None]*9)]
nm_base = [pct(v) for v in base_model.get('Net Margin:', [None]*9)]
nm_pess = [pct(v) for v in pess_model.get('Net Margin:', [None]*9)]
nm_opt  = [pct(v) for v in opt_model.get('Net Margin:',  [None]*9)]

income_statement = {
    'GrossProfit': {
        'Base':             gp_base,
        'Pessimistic':      gp_pess,
        'Optimistic':       gp_opt,
        'Weighted Average': gp_wa,
    },
    'GrossMargin': {
        'Base':             [r(a/b*100, 1) if b else None for a, b in zip(gp_base, rev_base)],
        'Pessimistic':      [r(a/b*100, 1) if b else None for a, b in zip(gp_pess, rev_pess)],
        'Optimistic':       [r(a/b*100, 1) if b else None for a, b in zip(gp_opt,  rev_opt)],
        'Weighted Average': [r(a/b*100, 1) if b else None for a, b in zip(gp_wa,   rev_wa)],
    },
    'OperatingMargin': {
        'Base':             om_base,
        'Pessimistic':      om_pess,
        'Optimistic':       om_opt,
        'Weighted Average': wa_row(om_base, om_pess, om_opt),
    },
    'NetIncome': {
        'Base':             ni_base,
        'Pessimistic':      ni_pess,
        'Optimistic':       ni_opt,
        'Weighted Average': ni_wa,
    },
    'NetMargin': {
        'Base':             nm_base,
        'Pessimistic':      nm_pess,
        'Optimistic':       nm_opt,
        'Weighted Average': wa_row(nm_base, nm_pess, nm_opt),
    },
}

# ── 4. DCF sensitivity tables ────────────────────────────────────────────────
dcf_ws   = wb['DCF']
dcf_rows = list(dcf_ws.iter_rows(values_only=True))

# WACC row is row index 6 (spreadsheet row 7)
wacc_header = dcf_rows[6]
# Implied share price WACC headers are at cols 23-31 (0-indexed)
wacc_values = [r(wacc_header[i], 5) for i in range(23, 32)]

# Base implied share prices: rows 8-16 → dcf_rows[7:16]
# Pessimistic:               rows 21-29 → dcf_rows[20:29]
# Optimistic:                rows 34-42 → dcf_rows[33:42]
# Each row: col 22 = TGR, cols 23-31 = implied prices for each WACC

def extract_sensitivity(row_start, row_end):
    result = []
    for row in dcf_rows[row_start:row_end]:
        tgr  = r(row[22], 4)
        prices = [r(row[i]) for i in range(23, 32)]
        result.append({'tgr': tgr, 'prices': prices})
    return result

# ── 4b. Evaluation sheet weighted-average sensitivity ────────────────────────
# WACC header: eval_rows[31] (spreadsheet row 32), cols 11-19
# Data rows:   eval_rows[32:41] (spreadsheet rows 33-41), TGR at col 11, prices at cols 12-20
eval_wacc_header = eval_rows[31]
wacc_values_wa = [r(eval_wacc_header[i], 5) for i in range(12, 21)]

def extract_eval_sensitivity():
    result = []
    for row in eval_rows[32:41]:
        tgr    = r(row[11], 4)
        prices = [r(row[i]) for i in range(12, 21)]
        result.append({'tgr': tgr, 'prices': prices})
    return result

dcf = {
    'waccLabels':     [f"{round(w*100, 3)}%" for w in wacc_values],
    'waccValues':     wacc_values,
    'waccLabelsWA':   [f"{round(w*100, 3)}%" for w in wacc_values_wa if w is not None],
    'waccValuesWA':   wacc_values_wa,
    'impliedPrices':  {
        'Base':        dcf_rows[20][5],   # single point (row 21, col 5)
        'Pessimistic': dcf_rows[20][6],
        'Optimistic':  dcf_rows[20][7],
    },
    'enterpriseValue': {
        'Base':        dcf_rows[18][5],   # row 19, col F
        'Pessimistic': dcf_rows[18][6],
        'Optimistic':  dcf_rows[18][7],
    },
    'sensitivity': {
        'Base':             extract_sensitivity(7,  16),
        'Pessimistic':      extract_sensitivity(20, 29),
        'Optimistic':       extract_sensitivity(33, 42),
        'Weighted Average': extract_eval_sensitivity(),
    },
    'assumptions': {
        'wacc':         r(dcf_rows[10][2], 4),   # row 11
        'tgr':          r(dcf_rows[13][2], 4),   # row 14
        'beta':         r(dcf_rows[9][2],  2),   # row 10
        'riskFreeRate': r(dcf_rows[7][2],  4),   # row 8
        'erp':          r(dcf_rows[8][2],  4),   # row 9
    },
}

# Round implied prices
for k in dcf['impliedPrices']:
    dcf['impliedPrices'][k] = r(dcf['impliedPrices'][k])
for k in dcf['enterpriseValue']:
    dcf['enterpriseValue'][k] = r(dcf['enterpriseValue'][k])

# Weighted Average implied price and EV (50% Base + 15% Pess + 35% Opt)
WEIGHTS = {'Base': 0.5, 'Pessimistic': 0.15, 'Optimistic': 0.35}
ip = dcf['impliedPrices']
ev = dcf['enterpriseValue']
if all(ip.get(s) is not None for s in WEIGHTS):
    ip['Weighted Average'] = r(sum(ip[s] * w for s, w in WEIGHTS.items()))
if all(ev.get(s) is not None for s in WEIGHTS):
    ev['Weighted Average'] = r(sum(ev[s] * w for s, w in WEIGHTS.items()))

# ── 5. Eval Comps table ──────────────────────────────────────────────────────
comp_ws   = wb['Eval Comps']
comp_rows = list(comp_ws.iter_rows(values_only=True))

COMP_COLS = [2, 3, 4, 5, 6, 7]   # Okta → Zscaler

def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime('%b %Y')
    return str(v) if v else None

comps = []
for ci in COMP_COLS:
    ev_rev  = r(comp_rows[14][ci], 2)   # row 15
    ev_ebda = r(comp_rows[15][ci], 2)   # row 16
    pe      = r(comp_rows[16][ci], 2)   # row 17
    comps.append({
        'name':     comp_rows[1][ci],           # row 2
        'ticker':   comp_rows[2][ci],           # row 3
        'fileDate': fmt_date(comp_rows[3][ci]), # row 4
        'price':    r(comp_rows[4][ci]),        # row 5
        'shares':   r(comp_rows[5][ci], 1),     # row 6  (diluted shares, M)
        'netDebt':  r(comp_rows[6][ci], 1),     # row 7  ($M)
        'revenue':  r(comp_rows[7][ci], 1),     # row 8  (TTM $M)
        'ebitda':   r(comp_rows[9][ci], 1),     # row 10 (TTM $M)
        'mktCap':   r(comp_rows[10][ci], 1),    # row 11 ($M)
        'ev':       r(comp_rows[11][ci], 1),    # row 12 ($M)
        'evRev':    None if (ev_rev  is None or ev_rev  <= 0) else ev_rev,
        'evEbitda': None if (ev_ebda is None or ev_ebda <= 0 or ev_ebda > 500) else ev_ebda,
        'pe':       None if (pe      is None or pe      <= 0) else pe,
    })

# ── 6. Consolidated Eval. football field ─────────────────────────────────────
consol_eval_ws   = wb['Consolidated Eval.']
consol_eval_rows = list(consol_eval_ws.iter_rows(values_only=True))

INCLUDE_TYPES = {'Implied Share Price', 'Attribute'}
football_field = []
for row in consol_eval_rows[1:]:          # skip header row
    if row[0] is None:
        continue
    if row[1] in INCLUDE_TYPES:
        football_field.append({
            'label':    row[0],
            'dataType': row[1],
            'low':      r(row[2]),
            'avg':      r(row[3]),
            'high':     r(row[4]),
        })

# ── 6. Assemble and write output ─────────────────────────────────────────────
output = {
    'generatedAt':   datetime.now().isoformat(),
    'currentPrice':  current_price,
    'years':         years,
    'metrics':       metrics,
    'incomeStatement': income_statement,
    'dcf':           dcf,
    'footballField': football_field,
    'comps':         comps,
}

os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
with open(OUTPUT_PATH, 'w') as f:
    json.dump(output, f, indent=2, default=str)

print(f"Done: {OUTPUT_PATH}")
print(f"  Current price: ${current_price}")
print(f"  DCF implied (Wtd Avg): ${dcf['impliedPrices'].get('Weighted Average')}")
print(f"  DCF implied (Base):    ${dcf['impliedPrices']['Base']}")
print(f"  DCF implied (Pess):    ${dcf['impliedPrices']['Pessimistic']}")
print(f"  DCF implied (Opt):     ${dcf['impliedPrices']['Optimistic']}")
print(f"  WA WACC labels: {dcf['waccLabelsWA']}")
print(f"  Years: {years}")
